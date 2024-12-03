import os
os.environ["CUDA_DEVICE_ORDER"] = "PCI_BUS_ID"
os.environ['CUDA_VISIBLE_DEVICES'] = '2'
import torch
from tqdm import tqdm
from prefetch_generator import BackgroundGenerator
import cv2
import logging
import sys
import openpyxl
import model as model
import utils
import numpy as np
from options.test_options import TestOptions
import natsort
import BatchDataReader

from pprint import pprint

def test_net(net,device):
    test_num = opt.test_ids[1] - opt.test_ids[0]
    test_dataset = BatchDataReader.CubeDataset(opt.dataroot, opt.test_ids, opt.data_size, opt.modality_filename,opt.label_filename,is_dataaug=False)
    test_loader = torch.utils.data.DataLoader(test_dataset, batch_size=1, shuffle=False, num_workers=1)
    test_results = os.path.join(opt.saveroot, 'test_results')
    BGR=np.zeros((opt.data_size[1],opt.data_size[2],3))

    with torch.no_grad():
        net.eval()
        val_miou_sum = 0
        val_acc_sum = 0                                                                     
        val_dice_sum = 0
        val_iou_sum = 0
        val_mdice_sum=0
        val_mdice_avg=0
        #test set
        row=1
        book = openpyxl.Workbook() #创建Excel
        book.create_sheet('sheet2') #创建sheet页
        table = book.active
        pbar = tqdm(enumerate(BackgroundGenerator(test_loader)), total=len(test_loader))
        for itr, (test_images, test_annotations,cubename) in pbar:
            test_images = test_images.to(device=device, dtype=torch.float32)
            pred= net(test_images)
            test_annotations_1 = np.squeeze(test_annotations).cpu().detach().numpy()
            test_annotations = test_annotations.cpu().detach().numpy()
            
            pred_argmax = torch.argmax(pred, dim=1)
            pred_argmax_1 = pred_argmax.cpu().detach().numpy()
            # print(cubename[0])
            # cv2.imwrite(os.path.join(test_results, cubename[0]), pred_argmax[0,0, :, :])
            pred_softmax = torch.nn.functional.softmax(pred, dim=1)
            pred_softmax = pred_softmax.cpu().detach().numpy()
            result= np.squeeze(pred_argmax).cpu().detach().numpy()
            cal_miou = utils.cal_miou(pred_argmax_1, test_annotations_1)
            val_miou_sum += cal_miou
            cavf_iou = utils.cal_cavfIOU(pred_argmax_1, test_annotations_1)
            val_iou_sum += cavf_iou
            perdice = utils.dice_coefficient_per_class(pred_argmax_1, test_annotations,[0,1,2,3,4])
            val_mdice_sum += perdice
            val_mdice_avg += np.mean(perdice)
            col=1
            table.cell(row,col,str(cubename))
            col+=1
            for one in cavf_iou:
                # print(one)
                table.cell(row,col,str(one))
                col+=1
            row+=1
            val_acc_sum += utils.cal_acc(result, test_annotations_1)
            val_dice_sum += utils.cal_Dice(result, test_annotations_1)
            # [0,0,:,:]是背景
            # [0,1,:,:]是毛细血管
            # [0,2,:,:]是红色
            # [0,3,:,:]是蓝色
            # [0,4,:,:]是FAZ
            BGR[:,:,0] = 53 * pred_softmax[0,0,:,:] + 143 * pred_softmax[0,1,:,:] + 28 * pred_softmax[0,2,:,:] + 186 * pred_softmax[0,3,:,:] + 106 * pred_softmax[0,4,:,:]
            BGR[:,:,1] = 32 * pred_softmax[0,0,:,:] + 165 * pred_softmax[0,1,:,:] + 25 * pred_softmax[0,2,:,:] + 131 * pred_softmax[0,3,:,:] + 217 * pred_softmax[0,4,:,:]
            BGR[:,:,2] = 15 * pred_softmax[0,0,:,:] + 171 * pred_softmax[0,1,:,:] + 215 * pred_softmax[0,2,:,:] + 43 * pred_softmax[0,3,:,:] + 166 * pred_softmax[0,4,:,:]
            # BGR[:,:,0] = 0 * pred_softmax[0,0,:,:] + 0 * pred_softmax[0,1,:,:] + 0 * pred_softmax[0,2,:,:] + 0 * pred_softmax[0,3,:,:] + 106 * pred_softmax[0,4,:,:]
            # BGR[:,:,1] = 0 * pred_softmax[0,0,:,:] + 0 * pred_softmax[0,1,:,:] + 0 * pred_softmax[0,2,:,:] + 0 * pred_softmax[0,3,:,:] + 217 * pred_softmax[0,4,:,:]
            # BGR[:,:,2] = 0 * pred_softmax[0,0,:,:] + 0 * pred_softmax[0,1,:,:] + 0 * pred_softmax[0,2,:,:] + 0 * pred_softmax[0,3,:,:] + 166 * pred_softmax[0,4,:,:]
            cv2.imwrite(os.path.join(opt.saveroot, 'test_visuals', cubename[0]), BGR)
        val_miou = val_miou_sum/test_num
        val_acc = val_acc_sum/test_num
        val_dice = val_dice_sum/test_num
        print("Valid_mIoU:{}".format(val_iou_sum/test_num))
        print("Valid_mDice:{}".format(val_mdice_sum / test_num))
        print("Valid_mIoU:{}".format(val_miou))
        print("Valid_mDice:{}".format(val_mdice_avg/test_num))
        print("val_acc:{}".format(val_acc))
        print("val_dice:{}".format(val_dice))
        book.save('writeExcel.xlsx')
        return val_miou
    
if __name__ == '__main__':
    torch.cuda.empty_cache()
    #setting logging
    logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
    #loading options
    opt = TestOptions().parse()
    #setting GPU
    
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    logging.info(f'Using device {device}')
    #loading network

    net = model.H2C_Net(in_channels=opt.in_channels, n_classes=opt.n_classes)
    #load trained model
    restore_path = '/home/image/PZY/Code/H2C_Net/logs/best_model'
    torch.cuda.empty_cache()
    if os.path.isfile(restore_path):
    # 如果是文件路径，获取其目录部分
        restore_path = os.path.dirname(restore_path)
        
    checkpoints=os.listdir(restore_path)
    test={}
    for checkpoint_path in tqdm(checkpoints):
        print("checkpoint_path:{}".format(checkpoint_path))  
        net.load_state_dict(
            torch.load(restore_path+'/'+checkpoint_path, map_location=device)
        )
        #input the model into GPU
        net.to(device=device)
        try:
            test_iou = test_net(net=net,device=device)
            test[checkpoint_path]=test_iou
        except KeyboardInterrupt:
            torch.save(net.state_dict(), 'INTERRUPTED.pth')
            logging.info('Saved interrupt')
            try:
                sys.exit(0)
            except SystemExit:
                os._exit(0)
    pprint(test)
